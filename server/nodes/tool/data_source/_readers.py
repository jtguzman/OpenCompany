"""Bounded, typed read tiers for the Data node.

Every reader here is synchronous (the node runs them via ``asyncio.to_thread``)
and self-bounding: tool results have NO downstream truncation before they are
serialized verbatim into an LLM message and the durable journal, so the caps
in this module are the only thing standing between a 20 MB spreadsheet and a
blown 2 MiB Temporal payload. ``bound_result`` is the final backstop that
trims any envelope over ``DATA_RESULT_MAX_BYTES``.

Binary content never leaves as bytes or base64 — the binary/image tiers
return references/metadata only (this deliberately fixes ``fileRead``'s
unbounded whole-file base64 fallback rather than inheriting it).
"""

from __future__ import annotations

import csv
import fnmatch
import hashlib
import io
import json
import mimetypes
import os
from pathlib import Path
from typing import Any, Optional

from services.plugin import NodeUserError

# Serialized-envelope ceiling: ~2/5 of the 512 KiB Temporal payload WARN
# threshold, leaving room for the JSON quoting the agent layer adds.
DATA_RESULT_MAX_BYTES = 200_000
TEXT_WINDOW_MAX_BYTES = 65_536
CSV_MAX_ROWS = 500
CSV_MAX_COLS = 100
CSV_MAX_CELL_CHARS = 2_000
JSON_MAX_SOURCE_BYTES = 5 * 1024 * 1024
JSON_MAX_DEPTH = 8
JSON_MAX_PRUNED_PATHS = 50
PDF_MAX_PAGES = 20
PDF_MAX_PAGE_CHARS = 20_000
XLSX_MAX_ROWS = 500
SEARCH_MAX_DEPTH = 12
SEARCH_MAX_VISITED = 5_000
SEARCH_MAX_RESULTS = 200

_TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".log", ".rst", ".ini", ".cfg", ".toml",
    ".yaml", ".yml", ".py", ".js", ".ts", ".tsx", ".jsx", ".css", ".sh",
    ".ps1", ".bat", ".sql", ".xml", ".env", ".jsonl", ".ndjson", ".srt",
    ".vtt", ".tex",
}


def detect_tier(path: Path, as_type: str = "auto") -> str:
    """Pick a read tier. An explicit ``as_type`` always wins."""
    if as_type and as_type != "auto":
        return as_type
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return "csv"
    if suffix == ".json":
        return "json"
    if suffix == ".pdf":
        return "pdf"
    if suffix in {".html", ".htm", ".xhtml"}:
        return "html"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    if suffix in _TEXT_EXTENSIONS:
        return "text"
    mime = mimetypes.guess_type(path.name)[0] or ""
    if mime.startswith("image/"):
        return "image"
    if mime.startswith("text/"):
        return "text"
    return "binary"


def _decode(
    raw: bytes, encoding: Optional[str]
) -> tuple[str, str, bool]:
    """Decode bytes -> (text, encoding_used, guessed)."""
    if encoding:
        try:
            return raw.decode(encoding), encoding, False
        except (LookupError, UnicodeDecodeError) as exc:
            raise NodeUserError(
                f"File does not decode as {encoding!r}: {exc}"
            ) from exc
    try:
        return raw.decode("utf-8"), "utf-8", False
    except UnicodeDecodeError:
        pass
    try:
        return raw.decode("utf-8-sig"), "utf-8-sig", True
    except UnicodeDecodeError:
        pass
    # latin-1 maps every byte; the flag tells the model this was a guess.
    return raw.decode("latin-1"), "latin-1", True


def _window_lines(
    text: str, offset: int, limit: int
) -> tuple[str, int, bool]:
    lines = text.splitlines()
    window = lines[offset : offset + limit]
    joined = "\n".join(window)
    truncated = offset + limit < len(lines)
    if len(joined.encode("utf-8", errors="replace")) > TEXT_WINDOW_MAX_BYTES:
        encoded = joined.encode("utf-8", errors="replace")[:TEXT_WINDOW_MAX_BYTES]
        joined = encoded.decode("utf-8", errors="replace")
        truncated = True
    return joined, len(lines), truncated


def read_text(
    path: Path, *, offset: int, limit: int, encoding: Optional[str]
) -> dict[str, Any]:
    text, used, guessed = _decode(path.read_bytes(), encoding)
    window, total, truncated = _window_lines(text, offset, limit)
    return {
        "type": "text",
        "text": window,
        "lines_total": total,
        "offset": offset,
        "limit": limit,
        "encoding": used,
        "encoding_guessed": guessed,
        "truncated": truncated,
    }


def read_csv(
    path: Path, *, offset: int, limit: int, encoding: Optional[str]
) -> dict[str, Any]:
    text, used, guessed = _decode(path.read_bytes(), encoding)
    sample = text[:16_384]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)

    def clip(cell: str) -> str:
        return cell if len(cell) <= CSV_MAX_CELL_CHARS else cell[:CSV_MAX_CELL_CHARS]

    columns: list[str] = []
    rows: list[list[str]] = []
    data_rows = 0
    cols_truncated = False
    capped_limit = min(limit, CSV_MAX_ROWS)
    for index, row in enumerate(reader):
        if index == 0:
            columns = [clip(cell) for cell in row[:CSV_MAX_COLS]]
            cols_truncated = cols_truncated or len(row) > CSV_MAX_COLS
            continue
        data_rows += 1
        position = index - 1
        if position < offset or len(rows) >= capped_limit:
            continue
        cols_truncated = cols_truncated or len(row) > CSV_MAX_COLS
        rows.append([clip(cell) for cell in row[:CSV_MAX_COLS]])
    return {
        "type": "csv",
        "columns": columns,
        "rows": rows,
        "rows_total": data_rows,
        "offset": offset,
        "delimiter": delimiter,
        "encoding": used,
        "encoding_guessed": guessed,
        "columns_truncated": cols_truncated,
        "truncated": offset + len(rows) < data_rows,
    }


def _prune_json(
    value: Any, depth: int, path: str, pruned: list[str]
) -> Any:
    if isinstance(value, dict):
        if depth >= JSON_MAX_DEPTH:
            if len(pruned) < JSON_MAX_PRUNED_PATHS:
                pruned.append(path or "$")
            return f"<pruned object: {len(value)} keys>"
        return {
            str(key): _prune_json(item, depth + 1, f"{path}.{key}", pruned)
            for key, item in value.items()
        }
    if isinstance(value, list):
        if depth >= JSON_MAX_DEPTH:
            if len(pruned) < JSON_MAX_PRUNED_PATHS:
                pruned.append(path or "$")
            return f"<pruned array: {len(value)} items>"
        return [
            _prune_json(item, depth + 1, f"{path}[{index}]", pruned)
            for index, item in enumerate(value)
        ]
    return value


def read_json(path: Path, *, encoding: Optional[str]) -> dict[str, Any]:
    raw = path.read_bytes()
    if len(raw) > JSON_MAX_SOURCE_BYTES:
        raise NodeUserError(
            f"JSON file is {len(raw):,} bytes (limit "
            f"{JSON_MAX_SOURCE_BYTES:,}). Read it as text with offset/limit "
            "instead: as_type='text'"
        )
    text, used, guessed = _decode(raw, encoding)
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise NodeUserError(f"File is not valid JSON: {exc}") from exc
    pruned: list[str] = []
    return {
        "type": "json",
        "data": _prune_json(data, 0, "", pruned),
        "pruned_paths": pruned,
        "encoding": used,
        "encoding_guessed": guessed,
        "truncated": bool(pruned),
    }


def read_pdf(path: Path, *, offset: int, limit: int) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise NodeUserError(
            "PDF reading needs the 'docs' extra: uv sync --extra docs "
            "(installs pypdf)"
        ) from None
    reader = PdfReader(str(path))
    pages_total = len(reader.pages)
    capped = min(limit, PDF_MAX_PAGES)
    pages: list[dict[str, Any]] = []
    for index in range(offset, min(offset + capped, pages_total)):
        text = reader.pages[index].extract_text() or ""
        pages.append(
            {
                "page": index + 1,
                "text": text[:PDF_MAX_PAGE_CHARS],
                "truncated": len(text) > PDF_MAX_PAGE_CHARS,
            }
        )
    return {
        "type": "pdf",
        "pages": pages,
        "pages_total": pages_total,
        "offset": offset,
        "truncated": offset + len(pages) < pages_total,
    }


def read_html(
    path: Path, *, offset: int, limit: int, encoding: Optional[str]
) -> dict[str, Any]:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise NodeUserError(
            "HTML extraction needs the 'docs' extra: uv sync --extra docs "
            "(installs beautifulsoup4). Or read the raw markup with "
            "as_type='text'"
        ) from None
    markup, used, guessed = _decode(path.read_bytes(), encoding)
    soup = BeautifulSoup(markup, "html.parser")
    title = soup.title.get_text(strip=True) if soup.title else None
    text = soup.get_text(separator="\n", strip=True)
    window, total, truncated = _window_lines(text, offset, limit)
    return {
        "type": "html",
        "title": title,
        "text": window,
        "lines_total": total,
        "offset": offset,
        "encoding": used,
        "encoding_guessed": guessed,
        "truncated": truncated,
    }


def read_xlsx(
    path: Path, *, sheet: Optional[str], offset: int, limit: int
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise NodeUserError(
            "XLSX reading needs openpyxl (pip install openpyxl)"
        ) from None
    try:
        workbook = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise NodeUserError(f"Cannot open spreadsheet: {exc}") from exc
    try:
        sheets = list(workbook.sheetnames)
        if sheet:
            if sheet not in sheets:
                raise NodeUserError(
                    f"No sheet named {sheet!r}. Sheets: {', '.join(sheets)}"
                )
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active
        capped = min(limit, XLSX_MAX_ROWS)

        def clip(cell: Any) -> Any:
            if isinstance(cell, str) and len(cell) > CSV_MAX_CELL_CHARS:
                return cell[:CSV_MAX_CELL_CHARS]
            return cell

        columns: list[Any] = []
        rows: list[list[Any]] = []
        cols_truncated = False
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            values = list(row)
            cols_truncated = cols_truncated or len(values) > CSV_MAX_COLS
            values = [clip(value) for value in values[:CSV_MAX_COLS]]
            if index == 0:
                columns = values
                continue
            position = index - 1
            if position < offset:
                continue
            if len(rows) >= capped:
                break
            rows.append(values)
        # max_row is header-inclusive metadata, cheap and close enough for
        # paging; an exact count would stream the whole sheet.
        rows_total = max(0, int(worksheet.max_row or 0) - 1)
        return {
            "type": "xlsx",
            "sheet": worksheet.title,
            "sheets": sheets,
            "columns": columns,
            "rows": rows,
            "rows_total": rows_total,
            "offset": offset,
            "columns_truncated": cols_truncated,
            "truncated": offset + len(rows) < rows_total,
        }
    finally:
        workbook.close()


def read_image_meta(path: Path) -> dict[str, Any]:
    """Header-only probe — dimensions and format, never pixels, no OCR."""
    try:
        from PIL import Image
    except ImportError:
        raise NodeUserError(
            "Image metadata needs Pillow (pip install pillow)"
        ) from None
    try:
        with Image.open(path) as image:
            return {
                "type": "image",
                "image": {
                    "width": image.width,
                    "height": image.height,
                    "format": (image.format or "").lower(),
                    "mode": image.mode,
                },
            }
    except Exception as exc:
        raise NodeUserError(f"Not a readable image: {exc}") from exc


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def walk_mount(
    root: Path, *, start_rel: str, pattern: str
) -> tuple[list[str], bool]:
    """Bounded name-glob walk of a mount. Returns (rel_paths, truncated).

    ``os.walk`` with ``followlinks=False`` (the default) so a symlinked
    directory can't pull in an unbounded external tree; symlinked *files*
    may appear in listings, but reading one later still goes through
    ``resolve_within``, which rejects escapes.
    """
    base = (root / start_rel) if start_rel else root
    base = base.resolve()
    results: list[str] = []
    visited = 0
    truncated = False
    base_depth = len(base.parts)
    for current, dirs, files in os.walk(base):
        depth = len(Path(current).parts) - base_depth
        if depth >= SEARCH_MAX_DEPTH:
            dirs[:] = []
        for name in files:
            visited += 1
            if visited > SEARCH_MAX_VISITED:
                truncated = True
                dirs[:] = []
                break
            if fnmatch.fnmatch(name.lower(), pattern.lower()):
                rel = os.path.relpath(
                    os.path.join(current, name), str(root)
                ).replace(os.sep, "/")
                results.append(rel)
                if len(results) >= SEARCH_MAX_RESULTS:
                    truncated = True
                    dirs[:] = []
                    break
        if truncated:
            break
    return results, truncated


_SHRINKABLE_LISTS = ("rows", "pages", "entries", "matches", "items")


def bound_result(result: dict[str, Any]) -> dict[str, Any]:
    """Final size backstop: trim known payload keys until the envelope fits."""

    def size(candidate: dict[str, Any]) -> int:
        return len(json.dumps(candidate, default=str).encode("utf-8"))

    if size(result) <= DATA_RESULT_MAX_BYTES:
        return result
    trimmed = dict(result)
    trimmed["truncated"] = True
    for key in _SHRINKABLE_LISTS:
        value = trimmed.get(key)
        while isinstance(value, list) and value and size(trimmed) > DATA_RESULT_MAX_BYTES:
            value = value[: max(1, len(value) // 2)]
            trimmed[key] = value
            if len(value) == 1 and size(trimmed) > DATA_RESULT_MAX_BYTES:
                break
    if size(trimmed) > DATA_RESULT_MAX_BYTES:
        for key in ("text", "data"):
            if key in trimmed and size(trimmed) > DATA_RESULT_MAX_BYTES:
                text = trimmed[key]
                if isinstance(text, str):
                    trimmed[key] = text[: TEXT_WINDOW_MAX_BYTES // 2]
                else:
                    trimmed[key] = "<payload removed: over result size cap>"
    return trimmed


__all__ = [
    "CSV_MAX_ROWS",
    "DATA_RESULT_MAX_BYTES",
    "PDF_MAX_PAGES",
    "SEARCH_MAX_RESULTS",
    "XLSX_MAX_ROWS",
    "bound_result",
    "detect_tier",
    "file_sha256",
    "read_csv",
    "read_html",
    "read_image_meta",
    "read_json",
    "read_pdf",
    "read_text",
    "read_xlsx",
    "walk_mount",
]
